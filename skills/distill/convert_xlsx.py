#!/usr/bin/env python3
"""Convert an Excel file to per-sheet CSV files.

Usage: convert_xlsx.py --input <path.xlsx> --output-dir <directory>

Output: One CSV per sheet at {basename}-{sheetname}.csv in output-dir.
"""

import argparse
import csv
import re
import sys
from pathlib import Path

try:
    from openpyxl import load_workbook
except ImportError:
    print("Error: openpyxl is not installed.", file=sys.stderr)
    print("Install with: pip install openpyxl==3.1.5", file=sys.stderr)
    sys.exit(1)

MAX_SHEETS = 20
NO_CACHE_SENTINEL = "#NO_CACHE"


def sanitize_sheetname(name: str) -> str:
    """Sanitize sheet name for use in filenames."""
    sanitized = re.sub(r"[^\w\s-]", "", name)
    sanitized = re.sub(r"\s+", "-", sanitized.strip())
    return sanitized.lower() or "sheet"


def convert_sheet_to_csv(
    ws_values, ws_formulas, output_path: Path
) -> dict:
    """Convert a single worksheet to CSV using two workbook views.

    ws_values: sheet from data_only=True workbook (cached values)
    ws_formulas: sheet from data_only=False workbook (formula detection)
    """
    total_cells = 0
    formula_cells = 0
    none_formula_cells = 0

    rows = []
    for row_vals, row_fmls in zip(ws_values.iter_rows(), ws_formulas.iter_rows()):
        csv_row = []
        for cell_val, cell_fml in zip(row_vals, row_fmls):
            total_cells += 1
            is_formula = isinstance(cell_fml.value, str) and cell_fml.value.startswith("=")

            if is_formula:
                formula_cells += 1
                if cell_val.value is None:
                    # Formula with no cached value
                    csv_row.append(NO_CACHE_SENTINEL)
                    none_formula_cells += 1
                else:
                    csv_row.append(str(cell_val.value))
            else:
                csv_row.append(str(cell_val.value) if cell_val.value is not None else "")

        rows.append(csv_row)

    with open(output_path, "w", newline="", encoding="utf-8") as f:
        writer = csv.writer(f)
        writer.writerows(rows)

    return {
        "total_cells": total_cells,
        "formula_cells": formula_cells,
        "none_formula_cells": none_formula_cells,
    }


def main():
    parser = argparse.ArgumentParser(description="Convert XLSX to CSV")
    parser.add_argument("--input", required=True, help="Path to input .xlsx file")
    parser.add_argument(
        "--output-dir", required=True, help="Directory for output CSV files"
    )
    args = parser.parse_args()

    input_path = Path(args.input)
    output_dir = Path(args.output_dir)

    if not input_path.exists():
        print(f"Error: Input file not found: {input_path}", file=sys.stderr)
        sys.exit(1)

    if not input_path.suffix.lower() == ".xlsx":
        print(
            f"Error: Expected .xlsx file, got: {input_path.suffix}", file=sys.stderr
        )
        sys.exit(1)

    if not output_dir.is_dir():
        print(f"Error: Output directory not found: {output_dir}", file=sys.stderr)
        sys.exit(1)

    basename = input_path.stem

    try:
        wb_values = load_workbook(str(input_path), data_only=True)
        wb_formulas = load_workbook(str(input_path), data_only=False)
    except Exception as e:
        print(f"Error opening {input_path}: {e}", file=sys.stderr)
        sys.exit(1)

    sheet_names = wb_values.sheetnames
    if len(sheet_names) > MAX_SHEETS:
        print(
            f"Warning: Workbook has {len(sheet_names)} sheets. "
            f"Converting only the first {MAX_SHEETS}.",
            file=sys.stderr,
        )
        sheet_names = sheet_names[:MAX_SHEETS]

    total_formula_cells = 0
    total_cells = 0
    total_none_formulas = 0
    output_files = []

    for sheet_name in sheet_names:
        ws_values = wb_values[sheet_name]
        ws_formulas = wb_formulas[sheet_name]
        sanitized = sanitize_sheetname(sheet_name)
        output_path = output_dir / f"{basename}-{sanitized}.csv"

        stats = convert_sheet_to_csv(ws_values, ws_formulas, output_path)
        total_cells += stats["total_cells"]
        total_formula_cells += stats["formula_cells"]
        total_none_formulas += stats["none_formula_cells"]
        output_files.append(str(output_path))

        print(f"Converted sheet '{sheet_name}' -> {output_path}")

    wb_values.close()
    wb_formulas.close()

    # Warnings
    if total_cells > 0 and total_formula_cells / total_cells > 0.3:
        print(
            f"Warning: {total_formula_cells}/{total_cells} cells "
            f"({total_formula_cells * 100 // total_cells}%) contain formulas. "
            f"Values may differ from displayed results.",
            file=sys.stderr,
        )

    if total_none_formulas > 0:
        print(
            f"Warning: {total_none_formulas} formula cells have no cached values "
            f"(written as {NO_CACHE_SENTINEL} in output). "
            f"Open and save in Excel/LibreOffice before distilling for accurate results.",
            file=sys.stderr,
        )

    print(f"Total: {len(output_files)} sheets converted")


if __name__ == "__main__":
    main()
